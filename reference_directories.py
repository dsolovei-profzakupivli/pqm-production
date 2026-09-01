import json
import re
import sqlite3
import threading
import urllib.parse
import urllib.request
from datetime import datetime

NAZK_URL = "https://corruptinfo.nazk.gov.ua/ep/1.0/corrupt/getAllData"
AMCU_PAGE = "https://amcu.gov.ua/napryami/oskarzhennya-publichnih-zakupivel/zvedeni-vidomosti-shchodo-spotvorennya-rezultativ-torgiv/zvedeni-vidomosti-shchodo-spotvorennia-rezultativ-torhiv-za-2026-rik"
LOCK = threading.Lock()
START_LOCK = threading.Lock()


def _now():
    return datetime.now().astimezone().isoformat(timespec="seconds")


def init_reference_tables(db_path):
    with sqlite3.connect(db_path, timeout=60) as con:
        con.execute("PRAGMA busy_timeout=60000")
        con.executescript("""
        CREATE TABLE IF NOT EXISTS reference_sync_state (
          source TEXT PRIMARY KEY, status TEXT NOT NULL DEFAULT 'idle', message TEXT DEFAULT '',
          updated_at TEXT DEFAULT '', source_updated_at TEXT DEFAULT '', row_count INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS nazk_registry (
          source_id TEXT PRIMARY KEY, punishment_type_code TEXT, punishment_type_name TEXT,
          entity_type_code TEXT, entity_type_name TEXT, last_name TEXT, first_name TEXT,
          patronymic TEXT, full_name TEXT, offense_id TEXT, offense_name TEXT, punishment TEXT,
          court_case_number TEXT, sentence_date TEXT, sentence_number TEXT, punishment_start TEXT,
          court_id TEXT, court_name TEXT, codex_articles TEXT, decision_url TEXT, raw_json TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_nazk_full_name ON nazk_registry(full_name);
        CREATE TABLE IF NOT EXISTS amcu_registry (
          row_key TEXT PRIMARY KEY, ordinal TEXT, division_no TEXT, sequence_no TEXT,
          decision_no TEXT, decision_date TEXT, authority TEXT, offender_name TEXT,
          offender_code TEXT, court_case_no TEXT, raw_json TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_amcu_code ON amcu_registry(offender_code);
        """)
        for source in ("nazk", "amcu"):
            con.execute("INSERT OR IGNORE INTO reference_sync_state(source) VALUES (?)", (source,))
        amcu_dates = [(normalized, key) for key, value in con.execute("SELECT row_key,decision_date FROM amcu_registry") if (normalized := _date_iso(value)) != (value or "")]
        nazk_dates = [(normalized, key) for key, value in con.execute("SELECT source_id,sentence_date FROM nazk_registry") if (normalized := _date_iso(value)) != (value or "")]
        if amcu_dates:
            con.executemany("UPDATE amcu_registry SET decision_date=? WHERE row_key=?", amcu_dates)
        if nazk_dates:
            con.executemany("UPDATE nazk_registry SET sentence_date=? WHERE source_id=?", nazk_dates)


def _state(db_path, source, status, message="", count=None, source_updated_at=None):
    fields = ["status=?", "message=?", "updated_at=?"]
    values = [status, message, _now()]
    if count is not None:
        fields.append("row_count=?"); values.append(int(count))
    if source_updated_at is not None:
        fields.append("source_updated_at=?"); values.append(source_updated_at)
    values.append(source)
    with sqlite3.connect(db_path) as con:
        con.execute(f"UPDATE reference_sync_state SET {','.join(fields)} WHERE source=?", values)


def reference_status(db_path):
    with sqlite3.connect(db_path) as con:
        con.row_factory = sqlite3.Row
        return {r["source"]: dict(r) for r in con.execute("SELECT * FROM reference_sync_state")}


def _fetch(url, timeout=900):
    # The AMCU portal rejects non-browser user agents even for public files.
    # Keep the request read-only, but identify it like a normal browser and
    # provide the public portal as referer for its static-object downloads.
    req = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/128 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*;q=0.8",
        "Accept-Language": "uk-UA,uk;q=0.9,en;q=0.7",
        "Referer": "https://amcu.gov.ua/",
    })
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return response.read()


def refresh_nazk(db_path):
    if not LOCK.acquire(blocking=False):
        return
    try:
        _state(db_path, "nazk", "running", "Завантаження реєстру НАЗК")
        payload = json.loads(_fetch(NAZK_URL).decode("utf-8-sig"))
        items = payload if isinstance(payload, list) else payload.get("data", payload.get("items", []))
        rows = []
        for i, item in enumerate(items):
            if not isinstance(item, dict): continue
            pt = item.get("punishmentType") or {}; et = item.get("entityType") or {}
            names = [item.get("indLastNameOnOffenseMoment"), item.get("indFirstNameOnOffenseMoment"), item.get("indPatronymicOnOffenseMoment")]
            full_name = " ".join(str(x).strip() for x in names if x and str(x).strip()).upper()
            sentence = str(item.get("sentenceNumber") or "").strip()
            decision_url = f"https://reyestr.court.gov.ua/Review/{sentence}" if sentence.isdigit() else ""
            articles = item.get("codexArticles")
            if not isinstance(articles, str): articles = json.dumps(articles, ensure_ascii=False) if articles is not None else ""
            rows.append((str(item.get("id") or i), str(pt.get("code") or ""), str(pt.get("name") or ""), str(et.get("code") or ""), str(et.get("name") or ""),
                *(str(x or "").strip() for x in names), full_name, str(item.get("offenseId") or ""), str(item.get("offenseName") or ""), str(item.get("punishment") or ""),
                str(item.get("courtCaseNumber") or ""), _date_iso(item.get("sentenceDate")), sentence, _date_iso(item.get("punishmentStart")), str(item.get("courtId") or ""),
                str(item.get("courtName") or ""), articles, decision_url, json.dumps(item, ensure_ascii=False)))
        with sqlite3.connect(db_path) as con:
            con.execute("BEGIN"); con.execute("DELETE FROM nazk_registry")
            con.executemany("INSERT INTO nazk_registry VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", rows)
        _state(db_path, "nazk", "ok", "Оновлено", len(rows), _now())
    except Exception as exc:
        _state(db_path, "nazk", "error", str(exc))
    finally:
        LOCK.release()


def start_reference_refresh(db_path, source, raw=None, filename=""):
    """Claim a reference refresh and defer heavy work until after HTTP 202 is flushed."""
    if source not in {"nazk", "amcu"}:
        raise ValueError("Невідомий довідник")
    with START_LOCK:
        state = reference_status(db_path).get(source, {})
        if state.get("status") == "running" or LOCK.locked():
            return False
        _state(db_path, source, "running", "Підготовка фонового оновлення")
        target = refresh_nazk if source == "nazk" else refresh_amcu
        args = (db_path,) if source == "nazk" else (db_path, raw, filename)
        timer = threading.Timer(0.2, target, args=args)
        timer.daemon = True
        timer.start()
    return True


def _cell(value):
    if value is None: return ""
    if isinstance(value, datetime): return value.date().isoformat()
    return str(value).strip()


def _date_iso(value):
    """Normalize registry dates to YYYY-MM-DD for reliable SQLite ordering."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    if not text:
        return ""
    for pattern in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y.%m.%d"):
        try:
            return datetime.strptime(text[:10], pattern).date().isoformat()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date().isoformat()
    except ValueError:
        return text


def _date_display(value):
    normalized = _date_iso(value)
    try:
        return datetime.strptime(normalized, "%Y-%m-%d").strftime("%d.%m.%Y")
    except ValueError:
        return normalized


def _find_col(headers, *needles):
    for idx, value in enumerate(headers):
        low = _cell(value).lower().replace("’", "'")
        if all(n in low for n in needles): return idx
    return None


def parse_amcu_xlsx(raw):
    from openpyxl import load_workbook
    import io
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    best = None
    for ws in wb.worksheets:
        values = list(ws.iter_rows(values_only=True))
        for pos, row in enumerate(values[:20]):
            headers = list(row)
            code_col = _find_col(headers, "ідентифікаційн")
            name_col = _find_col(headers, "суб'єкт", "поруш")
            if code_col is not None and name_col is not None:
                best = (values, pos, headers, code_col, name_col); break
        if best: break
    if not best: raise ValueError("У файлі АМКУ не знайдено очікувані заголовки")
    values, pos, headers, code_col, name_col = best
    date_col = _find_col(headers, "дата", "рішення")
    no_col = _find_col(headers, "№", "рішення")
    authority_col = _find_col(headers, "орган", "прийняв")
    court_col = _find_col(headers, "судов", "справ")
    rows_by_key = {}
    for n, row in enumerate(values[pos + 1:], 1):
        code = re.sub(r"\D", "", _cell(row[code_col] if code_col < len(row) else ""))
        name = _cell(row[name_col] if name_col < len(row) else "").upper()
        if not code and not name: continue
        # Official workbooks occasionally repeat column numbers or other
        # service rows inside the data range.  Ukrainian identifiers have at
        # least eight digits; foreign identifiers may be longer.
        if name.isdigit() or (code and len(code) < 8):
            continue
        get = lambda col: _cell(row[col]) if col is not None and col < len(row) else ""
        raw_row = {_cell(headers[i]) or f"column_{i+1}": _cell(v) for i, v in enumerate(row) if v is not None}
        decision_no = get(no_col)
        decision_date = _date_iso(row[date_col] if date_col is not None and date_col < len(row) else "")
        key = "|".join((code, decision_no, decision_date, name))
        # The official workbook can contain repeated entries.  Keep one row
        # for an identical business key so a duplicate cannot abort the whole
        # registry refresh with a UNIQUE constraint error.
        rows_by_key[key] = (key, str(n), "", "", decision_no, decision_date, get(authority_col), name, code, get(court_col), json.dumps(raw_row, ensure_ascii=False))
    return list(rows_by_key.values())


def discover_amcu_xlsx():
    html = _fetch(AMCU_PAGE, 90).decode("utf-8", "ignore")
    links = re.findall(r'href=["\']([^"\']+\.(?:xlsx|xls)(?:\?[^"\']*)?)["\']', html, re.I)
    if not links: raise ValueError("На сторінці АМКУ не знайдено Excel-файл")
    return urllib.parse.urljoin(AMCU_PAGE, links[0])


def refresh_amcu(db_path, raw=None, filename=""):
    if not LOCK.acquire(blocking=False): return
    try:
        _state(db_path, "amcu", "running", "Завантаження реєстру АМКУ")
        source = filename or discover_amcu_xlsx()
        rows = parse_amcu_xlsx(raw if raw is not None else _fetch(source, 180))
        with sqlite3.connect(db_path) as con:
            con.execute("BEGIN"); con.execute("DELETE FROM amcu_registry")
            con.executemany("INSERT OR REPLACE INTO amcu_registry VALUES (?,?,?,?,?,?,?,?,?,?,?)", rows)
        _state(db_path, "amcu", "ok", f"Оновлено з {source}", len(rows), _now())
    except Exception as exc:
        _state(db_path, "amcu", "error", str(exc))
    finally:
        LOCK.release()


def list_registry(db_path, source, query):
    q = (query.get("search", [""])[0] or "").strip()
    page = max(1, int(query.get("page", [1])[0])); size = min(200, max(1, int(query.get("size", [50])[0])))
    table = "nazk_registry" if source == "nazk" else "amcu_registry"
    cols = "full_name,offense_name,court_case_number,sentence_date,punishment_start,decision_url" if source == "nazk" else """offender_name,offender_code,decision_no,decision_date,authority,court_case_no,
      EXISTS(SELECT 1 FROM supplier_registry_summary srs WHERE srs.supplier_code=amcu_registry.offender_code) in_supplier_registry,
      COALESCE((SELECT MAX(srs.active_count) FROM supplier_registry_summary srs WHERE srs.supplier_code=amcu_registry.offender_code),0) active_qualifications,
      EXISTS(SELECT 1 FROM submissions s WHERE s.supplier_code=amcu_registry.offender_code) has_application"""
    conditions, params = [], []
    normalized_q = " ".join(re.sub(r"[’'`\-]+", " ", q.casefold()).split())
    if normalized_q:
        fields = ["full_name", "offense_name", "court_case_number"] if source == "nazk" else ["offender_name", "offender_code", "decision_no"]
        conditions.append("(" + " OR ".join(f"INSTR(NORMALIZE_SEARCH({f}),?)>0" for f in fields) + ")")
        params.extend([normalized_q] * len(fields))
    if source == "amcu":
        date_from = (query.get("date_from", [""])[0] or "").strip()
        date_to = (query.get("date_to", [""])[0] or "").strip()
        authority = (query.get("authority", [""])[0] or "").strip()
        supplier_scope = (query.get("supplier_scope", [""])[0] or "").strip()
        if date_from: conditions.append("decision_date>=?"); params.append(date_from)
        if date_to: conditions.append("decision_date<=?"); params.append(date_to)
        if authority: conditions.append("authority=?"); params.append(authority)
        if supplier_scope == "registered":
            conditions.append("EXISTS(SELECT 1 FROM supplier_registry_summary srs WHERE srs.supplier_code=amcu_registry.offender_code)")
        elif supplier_scope == "applicant":
            conditions.append("EXISTS(SELECT 1 FROM submissions s WHERE s.supplier_code=amcu_registry.offender_code)")
    else:
        date_from = (query.get("date_from", [""])[0] or "").strip()
        date_to = (query.get("date_to", [""])[0] or "").strip()
        court = (query.get("court", [""])[0] or "").strip()
        if date_from: conditions.append("sentence_date>=?"); params.append(date_from)
        if date_to: conditions.append("sentence_date<=?"); params.append(date_to)
        if court: conditions.append("court_name=?"); params.append(court)
    where = " WHERE " + " AND ".join(conditions) if conditions else ""
    order_by = "decision_date DESC, offender_name, row_key" if source == "amcu" else "sentence_date DESC, full_name, source_id"
    with sqlite3.connect(db_path) as con:
        con.row_factory = sqlite3.Row
        con.create_function("NORMALIZE_SEARCH", 1, lambda value: " ".join(
            re.sub(r"[’'`\-]+", " ", str(value or "").casefold()).split()
        ))
        total = con.execute(f"SELECT COUNT(*) FROM {table}{where}", params).fetchone()[0]
        rows = [dict(r) for r in con.execute(f"SELECT {cols} FROM {table}{where} ORDER BY {order_by} LIMIT ? OFFSET ?", params + [size, (page-1)*size])]
        date_field = "sentence_date" if source == "nazk" else "decision_date"
        for row in rows:
            row[date_field] = _date_display(row.get(date_field))
            if source == "nazk":
                row["punishment_start"] = _date_display(row.get("punishment_start"))
        authorities = [r[0] for r in con.execute("SELECT DISTINCT authority FROM amcu_registry WHERE authority<>'' ORDER BY authority")] if source == "amcu" else []
        courts = [r[0] for r in con.execute("SELECT DISTINCT court_name FROM nazk_registry WHERE court_name<>'' ORDER BY court_name")] if source == "nazk" else []
    return {"items": rows, "total": total, "page": page, "pages": max(1, (total+size-1)//size), "size": size, "authorities": authorities, "courts": courts}
